# -*- coding: utf-8 -*-
"""
SEGUNDA ETAPA - Transforma o CSV exportado do Sitrack em
<DATA>_listadoequipos_despositos.xlsx (abas Dados, Análise, Historico).

A linha "Depósito Trânsito a Base BR" (TIPO "B") entra como ULTIMA linha da
Análise e AGORA COMPOE o Total Geral / STD / CAMERAS (o SUBTOTAL inclui ela).
A aba Historico tem, para STD e CAMERAS: Total, Ativação, Retorno, "Tr. p. CI"
(tipo T), "Tr. p.Base" (tipo B = Base BR) e Outros. Total = A+R+Tr.CI+Tr.Base+O.
O historico.json guarda os totais COMPOSTOS (C.I. + Base BR).
"""

import os
import re
import sys
import glob
import json
import unicodedata
import datetime as dt
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Color
from openpyxl.utils import get_column_letter

SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
PASTA_RELAT  = os.path.join(SCRIPT_DIR, "relatorios")
PASTA_SAIDA  = os.path.join(PASTA_RELAT, "Template")
HIST_JSON    = os.path.join(PASTA_RELAT, "historico.json")
SUFIXO_SAIDA = "listadoequipos_despositos.xlsx"

COL_DESTINO  = "ÚLTIMO MOVIMIENTO EQUIPAMENTO (DESTINO)"
COL_FECHA    = "ÚLTIMO MOVIMIENTO EQUIPAMENTO (FECHA)"
COL_POSICAO  = "ÚLTIMA POSIÇÃO"
ROTULO_VAZIO = "Na"

VALOR_STD     = 150
VALOR_CAMERAS = 700
FILTRO_CLIENTE_PREFIXO = "C.I"

BASE_BR_NOME = "Depósito Trânsito a Base BR"
BASE_BR_TIPO = "B"

# "Depósito Retorno Sitrack" NAO entra na contagem total; aparece como linha
# separada (amarela) 3 linhas abaixo da Base BR, FORA do SUBTOTAL.
RETORNO_SITRACK_NOME = "Depósito Retorno Sitrack"

ORDEM_DADOS = [
    "EQUIPAMENTO", "MODELO", "CLIENTE (ID)", "CLIENTE (NOME)", "CLIENTE (ESTADO)",
    "NOME", "DOMINO",
    COL_POSICAO, COL_FECHA,
    "ÚLTIMO MOVIMIENTO EQUIPAMENTO (ORIGEN)", COL_DESTINO,
]
COLS_INT  = ["EQUIPAMENTO", "CLIENTE (ID)"]
COLS_DATA = [COL_POSICAO, COL_FECHA]
TIPOS = ["A", "R", "T", "O"]


def tipo_deposito(nome: str) -> str:
    s = unicodedata.normalize("NFKD", str(nome or "")).encode("ascii", "ignore").decode().lower()
    if "retorno" in s:
        return "R"
    if "ativacao" in s:
        return "A"
    if "transito" in s:
        return "T"
    return "O"


def categoria_std(modelo: str) -> bool:
    m = (modelo or "").strip()
    # STD145_RSTMini foi removido do somatorio de STD (passa a contar como "Outros")
    return m.startswith("STD") and "RSTMini" in m and m != "STD145_RSTMini"


def categoria_cameras(modelo: str) -> bool:
    return (modelo or "").strip().startswith("Streamax")


def to_int_ou_texto(v):
    s = (v or "").strip()
    if s == "":
        return None
    return int(s) if s.isdigit() else s


def achar_csv_mais_recente():
    arquivos = glob.glob(os.path.join(PASTA_RELAT, "*.csv"))
    return max(arquivos, key=os.path.getmtime) if arquivos else None


def extrair_data(caminho: str) -> str:
    m = re.search(r"(\d{4}-\d{2}-\d{2})", os.path.basename(caminho or ""))
    return m.group(1) if m else dt.date.today().isoformat()


def carregar_dados(caminho_csv: str) -> pd.DataFrame:
    df = pd.read_csv(caminho_csv, sep=";", encoding="latin-1", dtype=str, keep_default_na=False)
    df.columns = [c.strip() for c in df.columns]
    dest = df[COL_DESTINO].fillna("")
    partido = dest.str.split(" - ", n=1, expand=True)
    antes = partido[0].fillna("")
    depois = partido[1] if partido.shape[1] > 1 else pd.Series([""] * len(df))
    depois = depois.fillna("")
    sem_sep = ~dest.str.contains(" - ", regex=False)
    antes = antes.where(~sem_sep, "")
    depois = depois.where(~sem_sep, dest)
    df["NOME"] = depois.str.strip()
    df["DOMINO"] = antes.str.strip().map(to_int_ou_texto)
    for c in COLS_INT:
        df[c] = df[c].map(to_int_ou_texto)
    for c in COLS_DATA:
        df[c] = pd.to_datetime(df[c].replace("", None), format="%Y-%m-%d %H:%M:%S", errors="coerce")
    return df


def construir_analise(df: pd.DataFrame):
    cliente = df["CLIENTE (NOME)"].astype(str).str.strip().str.upper()
    dff = df[cliente.str.startswith(FILTRO_CLIENTE_PREFIXO.upper())]
    nome = dff["NOME"].replace("", ROTULO_VAZIO).fillna(ROTULO_VAZIO)
    modelo = dff["MODELO"].fillna("").replace("", "(sem modelo)")
    tab = pd.crosstab(nome, modelo)
    ci_models = list(tab.columns)

    bb = df[df["NOME"] == BASE_BR_NOME]
    bb_modelo = bb["MODELO"].fillna("").replace("", "(sem modelo)")
    bb_counts = bb_modelo.value_counts()

    modelos = sorted(set(ci_models) | set(bb_counts.index))
    tab = tab.reindex(columns=modelos, fill_value=0)
    tab["Total Geral"] = tab[modelos].sum(axis=1)
    std_cols = [m for m in modelos if categoria_std(m)]
    cam_cols = [m for m in modelos if categoria_cameras(m)]
    tab["STD"] = tab[std_cols].sum(axis=1) if std_cols else 0
    tab["CAMERAS"] = tab[cam_cols].sum(axis=1) if cam_cols else 0
    tab = tab.sort_values("Total Geral", ascending=False)

    # "Depósito Retorno Sitrack" sai da contagem total: vira linha separada (amarela)
    retorno = None
    if RETORNO_SITRACK_NOME in tab.index:
        rr = tab.loc[RETORNO_SITRACK_NOME]
        retorno = {"nome": RETORNO_SITRACK_NOME, "tipo": tipo_deposito(RETORNO_SITRACK_NOME),
                   "modelos": {m: int(rr[m]) for m in modelos},
                   "Total Geral": int(rr["Total Geral"]),
                   "STD": int(rr["STD"]), "CAMERAS": int(rr["CAMERAS"])}
        tab = tab.drop(index=RETORNO_SITRACK_NOME)

    tipos = {n: tipo_deposito(n) for n in tab.index}
    std_por_tipo = {t: 0 for t in TIPOS}
    cam_por_tipo = {t: 0 for t in TIPOS}
    for n, row in tab.iterrows():
        t = tipos[n]
        std_por_tipo[t] += int(row["STD"])
        cam_por_tipo[t] += int(row["CAMERAS"])

    bb_row = {m: int(bb_counts.get(m, 0)) for m in modelos}
    base_br = {
        "nome": BASE_BR_NOME, "tipo": BASE_BR_TIPO, "modelos": bb_row,
        "Total Geral": int(sum(bb_row.values())),
        "STD": int(sum(bb_row[m] for m in std_cols)),
        "CAMERAS": int(sum(bb_row[m] for m in cam_cols)),
    }
    resumo = {
        "Total Geral": int(tab["Total Geral"].sum()),
        "STD": int(tab["STD"].sum()),
        "CAMERAS": int(tab["CAMERAS"].sum()),
        "std_por_tipo": std_por_tipo,
        "cam_por_tipo": cam_por_tipo,
    }
    return tab, modelos, std_cols, cam_cols, tipos, resumo, base_br, retorno


AZUL    = PatternFill("solid", fgColor="FF1F4E78")
CINZA   = PatternFill("solid", fgColor="FFD9E1F2")
AMARELO = PatternFill("solid", fgColor="FFFFF2CC")
FILL_STD = PatternFill("solid", fgColor=Color(theme=2, tint=-0.499984740745262))
FILL_CAM = PatternFill("solid", fgColor=Color(theme=5, tint=-0.249977111117893))
BRANCO  = Font(color="FFFFFF", bold=True)
NEGRITO = Font(bold=True)
CENTRO  = Alignment(horizontal="center", vertical="center")
VERTICAL = Alignment(horizontal="center", vertical="bottom", textRotation=90)


def escrever_dados(ws, df: pd.DataFrame):
    df = df[ORDEM_DADOS]
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append([("" if (isinstance(v, float) and pd.isna(v)) or v is None else v) for v in row])
    for c in range(1, len(df.columns) + 1):
        cell = ws.cell(1, c); cell.fill = AZUL; cell.font = BRANCO; cell.alignment = CENTRO
    for nome_col in COLS_DATA:
        letra = get_column_letter(list(df.columns).index(nome_col) + 1)
        for r in range(2, ws.max_row + 1):
            ws[f"{letra}{r}"].number_format = "yyyy-mm-dd hh:mm:ss"
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{ws.max_row}"
    larguras = {"EQUIPAMENTO": 12, "MODELO": 18, "CLIENTE (ID)": 11,
                "CLIENTE (NOME)": 34, "CLIENTE (ESTADO)": 14, "NOME": 46,
                "DOMINO": 10, COL_POSICAO: 20, COL_FECHA: 20,
                "ÚLTIMO MOVIMIENTO EQUIPAMENTO (ORIGEN)": 46, COL_DESTINO: 46}
    for i, c in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = larguras.get(c, 16)


def escrever_analise(ws, tab, modelos, std_cols, cam_cols, tipos, base_br, retorno):
    cabec = ["DEPÓSITOS", "TIPO"] + modelos + ["Total Geral", "STD", "CAMERAS"]
    ws.append(cabec)
    ncols = len(cabec)
    ndep = len(tab)
    last_row = 2 + ndep + 1  # INCLUI a linha Base BR no subtotal

    linha2 = ["Total Geral", None]
    for c in range(3, ncols + 1):
        L = get_column_letter(c)
        linha2.append(f"=SUBTOTAL(9,{L}3:{L}{last_row})")
    ws.append(linha2)

    for nome_dep, linha in tab.iterrows():
        vals = [nome_dep, tipos.get(nome_dep, "O")]
        vals += [int(linha[m]) for m in modelos]
        vals += [int(linha["Total Geral"]), int(linha["STD"]), int(linha["CAMERAS"])]
        ws.append(vals)

    vals = [base_br["nome"], base_br["tipo"]]
    vals += [base_br["modelos"][m] for m in modelos]
    vals += [base_br["Total Geral"], base_br["STD"], base_br["CAMERAS"]]
    ws.append(vals)
    linha_base = ws.max_row

    # Retorno Sitrack (amarelo), 3 linhas abaixo da Base BR, FORA do subtotal
    linha_retorno = None
    if retorno:
        ws.append([]); ws.append([])
        rvals = [retorno["nome"], retorno["tipo"]]
        rvals += [retorno["modelos"][m] for m in modelos]
        rvals += [retorno["Total Geral"], retorno["STD"], retorno["CAMERAS"]]
        ws.append(rvals)
        linha_retorno = ws.max_row

    std_set, cam_set = set(std_cols), set(cam_cols)
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.font = BRANCO
        cell.alignment = VERTICAL
        rotulo = cabec[c - 1]
        if rotulo in std_set or rotulo == "STD":
            cell.fill = FILL_STD
        elif rotulo in cam_set or rotulo == "CAMERAS":
            cell.fill = FILL_CAM
        else:
            cell.fill = AZUL

    for c in range(1, ncols + 1):
        ws.cell(2, c).fill = CINZA; ws.cell(2, c).font = NEGRITO
    for c in range(1, ncols + 1):
        ws.cell(linha_base, c).fill = AMARELO; ws.cell(linha_base, c).font = NEGRITO
    if linha_retorno:
        for c in range(1, ncols + 1):
            ws.cell(linha_retorno, c).fill = AMARELO; ws.cell(linha_retorno, c).font = NEGRITO

    primeira_modelo, ultima_modelo = 3, 2 + len(modelos)
    for r in range(3, ws.max_row + 1):
        for c in range(primeira_modelo, ultima_modelo + 1):
            if ws.cell(r, c).value in (0, "0"):
                ws.cell(r, c).value = None

    ws.freeze_panes = "C3"
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 6
    for c in range(3, ncols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 6
    ws.row_dimensions[1].height = 150


def atualizar_historico(data_str, resumo, base_br):
    hist = []
    if os.path.exists(HIST_JSON):
        try:
            hist = json.load(open(HIST_JSON, encoding="utf-8"))
        except Exception:
            hist = []
    st = resumo["std_por_tipo"]; ct = resumo["cam_por_tipo"]
    entrada = {
        "data": data_str,
        "total": resumo["Total Geral"] + base_br["Total Geral"],
        "std": {"total": resumo["STD"] + base_br["STD"],
                "A": st["A"], "R": st["R"], "T": st["T"], "B": base_br["STD"], "O": st["O"]},
        "cameras": {"total": resumo["CAMERAS"] + base_br["CAMERAS"],
                    "A": ct["A"], "R": ct["R"], "T": ct["T"], "B": base_br["CAMERAS"], "O": ct["O"]},
    }
    hist = [h for h in hist if h.get("data") != data_str]
    hist.append(entrada)
    hist = sorted(hist, key=lambda h: h.get("data", ""))
    json.dump(hist, open(HIST_JSON, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    return hist


def _bloco_hist(v):
    if isinstance(v, dict):
        return [v.get("total", 0), v.get("A", ""), v.get("R", ""), v.get("T", ""), v.get("B", ""), v.get("O", "")]
    return [v or 0, "", "", "", "", ""]


def escrever_historico(ws, hist):
    ws.append([None, "STD", None, None, None, None, None,
               "CAMERAS", None, None, None, None, None])
    ws.append(["Data", "Total", "Ativação", "Retorno", "Tr. p. CI", "Tr. p.Base", "Outros",
               "Total", "Ativação", "Retorno", "Tr. p. CI", "Tr. p.Base", "Outros"])
    for h in hist:
        try:
            d = dt.datetime.strptime(h.get("data", ""), "%Y-%m-%d")
        except ValueError:
            d = h.get("data", "")
        ws.append([d] + _bloco_hist(h.get("std")) + _bloco_hist(h.get("cameras")))
        if isinstance(d, dt.datetime):
            ws.cell(ws.max_row, 1).number_format = "dd/mm/yyyy"
    ws["B1"].fill = FILL_STD; ws["B1"].font = BRANCO; ws["B1"].alignment = CENTRO
    ws["H1"].fill = FILL_CAM; ws["H1"].font = BRANCO; ws["H1"].alignment = CENTRO
    for c in range(1, 14):
        cell = ws.cell(2, c); cell.fill = AZUL; cell.font = BRANCO; cell.alignment = CENTRO
    ws.column_dimensions["A"].width = 12
    for c in range(2, 14):
        ws.column_dimensions[get_column_letter(c)].width = 11
    ws.freeze_panes = "A3"


def main():
    caminho_csv = sys.argv[1] if len(sys.argv) > 1 else achar_csv_mais_recente()
    if not caminho_csv or not os.path.exists(caminho_csv):
        print("[ERRO] CSV de entrada nao encontrado em", PASTA_RELAT)
        sys.exit(1)
    data_rel = extrair_data(caminho_csv)
    saida = sys.argv[2] if len(sys.argv) > 2 else os.path.join(PASTA_SAIDA, f"{data_rel}_{SUFIXO_SAIDA}")
    os.makedirs(os.path.dirname(saida), exist_ok=True)

    df = carregar_dados(caminho_csv)
    tab, modelos, std_cols, cam_cols, tipos, resumo, base_br, retorno = construir_analise(df)
    composto = resumo["Total Geral"] + base_br["Total Geral"]
    print("data", data_rel, "| C.I.", resumo["Total Geral"], "+ Base", base_br["Total Geral"],
          "= composto", composto, "| STD", resumo["STD"] + base_br["STD"],
          "| CAM", resumo["CAMERAS"] + base_br["CAMERAS"])

    hist = atualizar_historico(data_rel, resumo, base_br)

    wb = Workbook()
    ws_dados = wb.active
    ws_dados.title = "Dados"
    escrever_dados(ws_dados, df)
    escrever_analise(wb.create_sheet("Análise"), tab, modelos, std_cols, cam_cols, tipos, base_br, retorno)
    escrever_historico(wb.create_sheet("Historico"), hist)
    wb.save(saida)
    print("OK ->", saida)


if __name__ == "__main__":
    main()
